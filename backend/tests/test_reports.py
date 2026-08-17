"""Reports + Dashboard: stock summary, low stock, allocation report (JSON + xlsx export) and dashboard summary."""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.security import hash_password
from app.models.allocation import Allocation
from app.models.config import FiscalYear
from app.models.enums import AllocationStatus, AllocationType, ShipBaseType, StockTxnType, UserType
from app.models.inventory import Store
from app.models.item import Item, ItemCategory
from app.models.role import Role
from app.models.ship_base import ShipBase
from app.models.user import User
from app.services.stock_ledger import apply_stock_movement
from app.utils.export import rows_to_xlsx

API = "/api/v1"
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest_asyncio.fixture()
async def seeded(db_session):
    """Two stores, two categories, three items, two ship/bases, stocks (two low) and four allocations."""
    db = db_session
    cat_a = ItemCategory(code="CAT-A", name="Ropes")
    cat_b = ItemCategory(code="CAT-B", name="Radios")
    db.add_all([cat_a, cat_b])
    await db.flush()
    rope = Item(code="IT-001", name="Mooring Rope", category_id=cat_a.id)
    radio = Item(code="IT-002", name="VHF Radio", category_id=cat_b.id)
    cable = Item(code="IT-003", name="Coax Cable", category_id=cat_b.id)
    db.add_all([rope, radio, cable])
    central = Store(code="ST-01", name="Central Store")
    depot = Store(code="ST-02", name="Chattogram Depot")
    db.add_all([central, depot])
    frigate = ShipBase(code="SB-01", name="BNS Bangabandhu", type=ShipBaseType.SHIP)
    base = ShipBase(code="SB-02", name="BNS Issa Khan", type=ShipBaseType.BASE)
    db.add_all([frigate, base])
    await db.flush()

    await apply_stock_movement(
        db,
        store_id=central.id,
        item_id=rope.id,
        quantity_delta=100,
        txn_type=StockTxnType.OPENING,
        user_id=None,
        low_stock_threshold=Decimal("10"),
    )
    await apply_stock_movement(
        db,
        store_id=central.id,
        item_id=radio.id,
        quantity_delta=3,
        txn_type=StockTxnType.OPENING,
        user_id=None,
        low_stock_threshold=Decimal("5"),
    )  # LOW
    await apply_stock_movement(
        db,
        store_id=depot.id,
        item_id=cable.id,
        quantity_delta=0,
        txn_type=StockTxnType.OPENING,
        user_id=None,
        low_stock_threshold=Decimal("1"),
    )  # LOW
    await apply_stock_movement(
        db, store_id=depot.id, item_id=rope.id, quantity_delta=40, txn_type=StockTxnType.OPENING, user_id=None
    )  # no threshold -> never low

    fys = (await db.execute(select(FiscalYear).order_by(FiscalYear.start_date))).scalars().all()
    fy1, fy2 = fys[0], fys[1]
    admin = (await db.execute(select(User).where(User.username == "admin"))).scalars().first()
    allocs = [
        Allocation(
            code="AL-1",
            allocation_type=AllocationType.ALLOCATION,
            fiscal_year_id=fy1.id,
            allocation_date=date(2025, 1, 10),
            store_id=central.id,
            item_id=rope.id,
            ship_base_id=frigate.id,
            quantity=10,
            status=AllocationStatus.APPROVED,
            approved_by_id=admin.id,
        ),
        Allocation(
            code="AL-2",
            allocation_type=AllocationType.SANCTION,
            fiscal_year_id=fy1.id,
            allocation_date=date(2025, 2, 10),
            store_id=central.id,
            item_id=radio.id,
            ship_base_id=frigate.id,
            quantity=2,
            status=AllocationStatus.PENDING,
        ),
        Allocation(
            code="AL-3",
            allocation_type=AllocationType.ALLOCATION,
            fiscal_year_id=fy2.id,
            allocation_date=date(2025, 8, 1),
            store_id=depot.id,
            item_id=cable.id,
            ship_base_id=base.id,
            quantity=5,
            status=AllocationStatus.SENT_BACK,
        ),
        Allocation(
            code="AL-4",
            allocation_type=AllocationType.ALLOCATION,
            fiscal_year_id=fy2.id,
            allocation_date=date(2025, 9, 1),
            store_id=depot.id,
            item_id=rope.id,
            ship_base_id=base.id,
            quantity=7,
            status=AllocationStatus.PENDING,
        ),
    ]
    db.add_all(allocs)
    await db.flush()

    # a ship/base user bound to `base` (scoped views)
    role = (await db.execute(select(Role).where(Role.name == "Ship/Base User"))).scalars().first()
    sb_user = User(
        user_type=UserType.SHIP_BASE_USER,
        username="sbuser",
        full_name="Ship User",
        hashed_password=hash_password("Ship@12345"),
        role_id=role.id if role else None,
        ship_base_id=base.id,
    )
    db.add(sb_user)
    await db.commit()
    return {
        "central": central.id,
        "depot": depot.id,
        "rope": rope.id,
        "radio": radio.id,
        "cable": cable.id,
        "cat_a": cat_a.id,
        "cat_b": cat_b.id,
        "frigate": frigate.id,
        "base": base.id,
        "fy1": fy1.id,
        "fy2": fy2.id,
    }


def _xlsx_rows(content: bytes) -> list[list]:
    ws = load_workbook(io.BytesIO(content)).active
    return [list(r) for r in ws.iter_rows(values_only=True)]


@pytest.mark.asyncio
async def test_reports_require_auth(client: AsyncClient):
    assert (await client.get(f"{API}/reports/stock-summary")).status_code == 401
    assert (await client.get(f"{API}/reports/allocations")).status_code == 401
    assert (await client.get(f"{API}/dashboard/summary")).status_code == 401


@pytest.mark.asyncio
async def test_stock_summary_filters_sort_and_export(client: AsyncClient, admin_headers, seeded):
    h = admin_headers
    r = await client.get(f"{API}/reports/stock-summary", headers=h)
    assert r.status_code == 200, r.text
    j = r.json()
    assert j["total"] == 4 and len(j["items"]) == 4
    row = next(x for x in j["items"] if x["item"]["code"] == "IT-002")
    assert row["store"]["name"] == "Central Store"
    assert row["item"]["category"]["name"] == "Radios"
    assert (
        row["is_low"] is True and Decimal(row["quantity"]) == 3 and Decimal(row["low_stock_threshold"]) == 5
    )
    assert row["last_updated"]
    rope_depot = next(
        x for x in j["items"] if x["item"]["code"] == "IT-001" and x["store"]["code"] == "ST-02"
    )
    assert rope_depot["is_low"] is False and rope_depot["low_stock_threshold"] is None

    # explicit filters
    r = await client.get(f"{API}/reports/stock-summary", params={"store_id": seeded["depot"]}, headers=h)
    assert r.json()["total"] == 2
    r = await client.get(f"{API}/reports/stock-summary", params={"item_id": seeded["rope"]}, headers=h)
    assert r.json()["total"] == 2
    r = await client.get(f"{API}/reports/stock-summary", params={"category_id": seeded["cat_b"]}, headers=h)
    assert r.json()["total"] == 2
    r = await client.get(f"{API}/reports/stock-summary", params={"low_only": "true"}, headers=h)
    assert r.json()["total"] == 2 and all(x["is_low"] for x in r.json()["items"])
    # column filter row + q + sort + pagination
    r = await client.get(f"{API}/reports/stock-summary", params={"filter": "store:depot"}, headers=h)
    assert r.json()["total"] == 2
    r = await client.get(f"{API}/reports/stock-summary", params={"q": "vhf"}, headers=h)
    assert r.json()["total"] == 1
    r = await client.get(
        f"{API}/reports/stock-summary", params={"sort": "quantity:desc", "page_size": 2}, headers=h
    )
    j = r.json()
    assert j["pages"] == 2 and Decimal(j["items"][0]["quantity"]) == 100
    assert (
        await client.get(f"{API}/reports/stock-summary", params={"sort": "nope:asc"}, headers=h)
    ).status_code == 422
    assert (
        await client.get(f"{API}/reports/stock-summary", params={"filter": "nope:x"}, headers=h)
    ).status_code == 422

    # xlsx export honours filters
    r = await client.get(
        f"{API}/reports/stock-summary", params={"export": "xlsx", "store_id": seeded["depot"]}, headers=h
    )
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(XLSX)
    assert 'filename="cims_stock_summary_' in r.headers["content-disposition"]
    rows = _xlsx_rows(r.content)
    assert rows[0][:4] == ["Store Code", "Store", "Item ID", "Item Name"]
    assert len(rows) == 3 and {row[0] for row in rows[1:]} == {"ST-02"}
    assert (
        await client.get(f"{API}/reports/stock-summary", params={"export": "csv"}, headers=h)
    ).status_code == 422


@pytest.mark.asyncio
async def test_low_stock_report(client: AsyncClient, admin_headers, seeded):
    h = admin_headers
    r = await client.get(f"{API}/reports/low-stock", headers=h)
    assert r.status_code == 200
    j = r.json()
    assert j["total"] == 2
    assert {x["item"]["code"] for x in j["items"]} == {"IT-002", "IT-003"}
    assert all(x["is_low"] for x in j["items"])
    r = await client.get(f"{API}/reports/low-stock", params={"store_id": seeded["central"]}, headers=h)
    assert r.json()["total"] == 1
    r = await client.get(f"{API}/reports/low-stock", params={"export": "xlsx"}, headers=h)
    assert r.status_code == 200 and r.headers["content-type"].startswith(XLSX)
    assert "cims_low_stock_" in r.headers["content-disposition"]
    assert len(_xlsx_rows(r.content)) == 3


@pytest.mark.asyncio
async def test_allocation_report_filters_scope_and_export(client: AsyncClient, admin_headers, seeded):
    h = admin_headers
    r = await client.get(f"{API}/reports/allocations", headers=h)
    assert r.status_code == 200, r.text
    j = r.json()
    assert j["total"] == 4
    assert j["items"][0]["code"] == "AL-4"  # default sort date:desc
    row = next(x for x in j["items"] if x["code"] == "AL-1")
    assert row["type"] == "allocation" and row["status"] == "approved"
    assert row["fiscal_year"]["name"] and row["store"]["name"] == "Central Store"
    assert row["item"]["code"] == "IT-001" and row["ship_base"]["name"] == "BNS Bangabandhu"
    assert row["approved_by"]["username"] == "admin"
    assert Decimal(row["quantity"]) == 10

    async def total(**params):
        rr = await client.get(f"{API}/reports/allocations", params=params, headers=h)
        assert rr.status_code == 200, rr.text
        return rr.json()["total"]

    assert await total(fiscal_year_id=seeded["fy1"]) == 2
    assert await total(ship_base_id=seeded["base"]) == 2
    assert await total(store_id=seeded["central"]) == 2
    assert await total(item_id=seeded["rope"]) == 2
    assert await total(status="pending") == 2
    assert await total(type="sanction") == 1
    assert await total(date_from="2025-08-01") == 2
    assert await total(date_from="2025-02-01", date_to="2025-08-31") == 2
    assert await total(filter="ship_base:issa") == 2
    assert await total(filter="status:approved") == 1
    assert await total(q="AL-3") == 1
    assert (
        await client.get(f"{API}/reports/allocations", params={"status": "bogus"}, headers=h)
    ).status_code == 422
    r = await client.get(f"{API}/reports/allocations", params={"sort": "quantity:desc"}, headers=h)
    assert r.json()["items"][0]["code"] == "AL-1"

    r = await client.get(
        f"{API}/reports/allocations", params={"export": "xlsx", "status": "pending"}, headers=h
    )
    assert r.status_code == 200 and r.headers["content-type"].startswith(XLSX)
    assert "cims_allocations_" in r.headers["content-disposition"]
    rows = _xlsx_rows(r.content)
    assert rows[0][0] == "ID" and rows[0][9] == "Status"
    assert len(rows) == 3 and {row[9] for row in rows[1:]} == {"Pending"}

    # ship/base user only sees own ship/base rows (when its role grants report.list; otherwise 403)
    r = await client.post(f"{API}/auth/login", json={"identifier": "sbuser", "password": "Ship@12345"})
    assert r.status_code == 200
    sb_h = {"Authorization": f"Bearer {r.json()['access_token']}"}
    r = await client.get(f"{API}/reports/allocations", headers=sb_h)
    if r.status_code == 200:
        assert r.json()["total"] == 2 and {x["ship_base"]["code"] for x in r.json()["items"]} == {"SB-02"}
    else:
        assert r.status_code == 403


@pytest.mark.asyncio
async def test_dashboard_summary(client: AsyncClient, admin_headers, seeded):
    r = await client.get(f"{API}/dashboard/summary", headers=admin_headers)
    assert r.status_code == 200, r.text
    d = r.json()
    c = d["counts"]
    assert c["items"] == 3 and c["ship_bases"] == 2 and c["stores"] == 2 and c["users"] == 2
    assert (
        c["allocations_pending"] == 2 and c["allocations_approved"] == 1 and c["allocations_sent_back"] == 1
    )
    assert c["low_stock_items"] == 2

    by_status = {x["status"]: x["count"] for x in d["allocations_by_status"]}
    assert by_status == {"pending": 2, "approved": 1, "sent_back": 1, "cancelled": 0}

    fy = d["allocations_by_fiscal_year"]
    assert len(fy) == 2
    assert fy[0]["allocation"] == 1 and fy[0]["sanction"] == 1 and Decimal(fy[0]["total_qty"]) == 12
    assert fy[1]["allocation"] == 2 and fy[1]["sanction"] == 0 and Decimal(fy[1]["total_qty"]) == 12

    sb = {x["ship_base"]: x for x in d["allocations_by_ship_base"]}
    assert sb["BNS Bangabandhu"]["count"] == 2 and Decimal(sb["BNS Bangabandhu"]["qty"]) == 12
    assert sb["BNS Issa Khan"]["count"] == 2

    assert {x["category"]: x["count"] for x in d["items_by_category"]} == {"Radios": 2, "Ropes": 1}
    st = {x["store"]: x for x in d["stock_by_store"]}
    assert st["Central Store"]["items"] == 2 and Decimal(st["Central Store"]["total_qty"]) == 103
    assert st["Chattogram Depot"]["items"] == 2 and Decimal(st["Chattogram Depot"]["total_qty"]) == 40

    assert len(d["recent_allocations"]) == 4
    assert d["recent_allocations"][0]["code"] == "AL-4"
    assert {
        "id",
        "code",
        "type",
        "fiscal_year",
        "date",
        "store",
        "item",
        "ship_base",
        "quantity",
        "status",
    } <= set(d["recent_allocations"][0])
    assert len(d["low_stock"]) == 2 and all(x["is_low"] for x in d["low_stock"])
    assert d["low_stock"][0]["item"]["code"] == "IT-002"  # biggest shortfall first (3-5 = -2 vs 0-1 = -1)


def test_rows_to_xlsx_unit():
    content = rows_to_xlsx(["A", "Long header name"], [[1, "x"], [Decimal("2.5"), None]])
    ws = load_workbook(io.BytesIO(content)).active
    assert ws["A1"].value == "A" and ws["A1"].font.bold
    assert ws["A3"].value == 2.5 and ws["B3"].value in ("", None)
    assert ws.column_dimensions["B"].width >= len("Long header name")
