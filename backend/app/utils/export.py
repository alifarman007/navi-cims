"""Excel export helpers (openpyxl): generic rows -> xlsx bytes, and a StreamingResponse wrapper.

xlsx_response(headers=["Store", "Item"], rows=[["Central", "Rope"]], report="stock_summary")
-> attachment cims_stock_summary_2026-08-17.xlsx
"""

from __future__ import annotations

import io
from collections.abc import Iterable, Sequence
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_HEADER_FILL = PatternFill(start_color="E3E8FF", end_color="E3E8FF", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="1C3586")


def _cell_value(v: Any) -> Any:
    """Coerce python values into something openpyxl accepts."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v.replace(tzinfo=None) if v.tzinfo else v
    if isinstance(v, date):
        return v
    if hasattr(v, "value"):  # enums
        return str(v.value).replace("_", " ").title()
    if isinstance(v, (int, float, str)):
        return v
    return str(v)


def rows_to_xlsx(headers: Sequence[str], rows: Iterable[Sequence[Any]], sheet_name: str = "Report") -> bytes:
    """Build a workbook with a bold header row, frozen pane and autosized columns; return the bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Report"

    ws.append([str(h) for h in headers])
    for c in ws[1]:
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    widths = [len(str(h)) for h in headers]
    for row in rows:
        values = [_cell_value(v) for v in row]
        ws.append(values)
        for i, v in enumerate(values):
            if i < len(widths):
                widths[i] = max(widths[i], min(len(str(v)), 60))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w + 2

    for row in ws.iter_rows(min_row=2):
        for c in row:
            if isinstance(c.value, datetime):
                c.number_format = "yyyy-mm-dd hh:mm"
            elif isinstance(c.value, date):
                c.number_format = "yyyy-mm-dd"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(report: str, on: date | None = None) -> str:
    return f"cims_{report}_{(on or date.today()).isoformat()}.xlsx"


def xlsx_response(headers: Sequence[str], rows: Iterable[Sequence[Any]], report: str) -> StreamingResponse:
    """StreamingResponse for an xlsx attachment named cims_<report>_<date>.xlsx."""
    data = rows_to_xlsx(headers, rows, sheet_name=report.replace("_", " ").title())
    filename = export_filename(report)
    return StreamingResponse(
        io.BytesIO(data),
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
